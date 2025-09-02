from openpyxl import load_workbook
from openpyxl import Workbook
import re

# Load the Excel file
input_file = r"C:\\Users\\Admin\\Documents\\names\\All_Names.xlsx"
wb = load_workbook(filename=input_file)
sheet = wb.active

# Create separate Workbooks for English letters and non-English letters data
english_wb = Workbook()
non_english_wb = Workbook()

english_sheet = english_wb.active
non_english_sheet = non_english_wb.active

# Assuming you want to check the second column (column 'B')
column_index = 2  # Change this index to the desired column (1 for column 'A', 2 for 'B', etc.)

# Iterate through the rows and separate the data based on a particular column
for row in sheet.iter_rows(values_only=True):
    cell_value = row[column_index - 1]  # Adjusting to Python's 0-based index

    # Using regex to check if the cell starts with English letters
    if re.match(r'^[a-zA-Z]', str(cell_value)):
        english_sheet.append(row)
    else:
        non_english_sheet.append(row)

# Save the separated data into separate Excel files
english_wb.save(filename="english_data.xlsx")
non_english_wb.save(filename="non_english_data.xlsx")

